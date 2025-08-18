import csv
import os
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from graph_profit import get_stats, lire_trades_sql, filtrer_trades

# 🎨 Classe PDF avec header/footer personnalisés
class PDF(FPDF):
    def header(self):
        if os.path.exists("assets/logo.png"):
            self.image("assets/logo.png", 10, 8, 33)
        self.set_font("Arial", "B", 16)
        self.cell(0, 10, "Rapport IA - BrainTrader-X1", ln=True, align="C")
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        date_gen = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.cell(0, 10, f"Page {self.page_no()} - Généré le {date_gen}", align="C")

# 📄 Rapport PDF cockpit IA
def export_cockpit_pdf():
    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "📊 Rapport Cockpit IA", ln=True, align="C")

    # 🧠 Statistiques
    stats = get_stats()
    pdf.set_font("Arial", "", 12)
    pdf.ln(8)
    pdf.cell(0, 10, f"Total trades : {stats['total']}", ln=True)
    pdf.cell(0, 10, f"Gagnants : {stats['gagnants']} | Perdants : {stats['perdants']}", ln=True)
    pdf.ln(8)

    # 📈 Graphes
    dossier_static = os.path.join(os.path.dirname(__file__), "static")
    graph1 = os.path.join(dossier_static, "capital_graph.png")
    graph2 = os.path.join(dossier_static, "strategie_graph.png")
    if os.path.exists(graph1): pdf.image(graph1, w=180); pdf.ln(6)
    if os.path.exists(graph2): pdf.image(graph2, w=180); pdf.ln(10)

    # 📋 Historique
    trades = lire_trades_sql()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "📋 Historique des trades", ln=True)
    pdf.set_font("Arial", "", 10)
    for t in trades[:10]:
        pdf.cell(0, 6, f"{t['date']} {t['heure']} - {t['symbole']} - {t['profit']}€ - Strat: {t['strategie']}", ln=True)

    os.makedirs("static", exist_ok=True)
    pdf.output(os.path.join("static", "rapport_ia.pdf"))

# 📥 Export CSV des trades filtrés
def export_csv_trades(filtres):
    trades = filtrer_trades(filtres)
    os.makedirs("static", exist_ok=True)
    chemin = os.path.join("static", "trades_export.csv")
    with open(chemin, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "Heure", "Symbole", "Volume", "Profit (€)", "Stratégie"])
        for t in trades:
            writer.writerow([t["date"], t["heure"], t["symbole"], t["volume"], t["profit"], t["strategie"]])

# 📊 Export Excel des trades filtrés + stats IA
def export_excel_trades(filtres):
    trades = filtrer_trades(filtres)
    stats = get_stats()
    wb = Workbook()
    ws_trades = wb.active
    ws_trades.title = "Trades IA"

    colonnes = ["Date", "Heure", "Symbole", "Volume", "Profit (€)", "Stratégie"]
    ws_trades.append(colonnes)
    for t in trades:
        ws_trades.append([t["date"], t["heure"], t["symbole"], t["volume"], t["profit"], t["strategie"]])

    for col in ws_trades.iter_cols(min_row=2, min_col=5, max_col=5):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.font = Font(color="FF0000") if cell.value < 0 else Font(color="008800")

    ws_stats = wb.create_sheet("Stats IA")
    ws_stats.append(["📊 Statistiques globales"])
    ws_stats.append(["Total trades", stats["total"]])
    ws_stats.append(["Gagnants", stats["gagnants"]])
    ws_stats.append(["Perdants", stats["perdants"]])
    ws_stats.append([])
    ws_stats.append(["Profit total par symbole"])
    for s in stats["symboles"]:
        ws_stats.append([s[0], s[1]])
    ws_stats.append([])
    ws_stats.append(["Profit moyen par stratégie"])
    for s in stats["strategies"]:
        ws_stats.append([s[0], s[1]])

    os.makedirs("static", exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    chemin = os.path.join("static", f"rapport_ia_{date_str}.xlsx")
    wb.save(chemin)
    print(f"✅ Fichier Excel généré : {chemin}")